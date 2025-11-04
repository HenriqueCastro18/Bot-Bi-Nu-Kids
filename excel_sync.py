# excel_sync.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import database
import os
import json
import re 
from datetime import datetime

ARQUIVO_EXCEL = "gestão de custo.xlsx"
PLANILHA_ABA = "Vendas"

# --- (FUNÇÃO HELPER) ---
def formatar_cpf(cpf_numeros: str) -> str:
    """Formata uma string de 11 dígitos de CPF para XXX.XXX.XXX-XX."""
    if not cpf_numeros or len(cpf_numeros) != 11 or not cpf_numeros.isdigit():
        return cpf_numeros 
    return f"{cpf_numeros[0:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:11]}"

# --- (FUNÇÃO HELPER) ---
def formatar_itens_para_excel(itens_json: str) -> str:
    """Converte o JSON de itens em uma string legível."""
    try:
        lista_itens = json.loads(itens_json)
        itens_formatados = []
        
        for item in lista_itens:
            nome = item.get('nome', 'Item desconhecido')
            preco = item.get('preco', 'R$ 0,00')
            
            itens_formatados.append(f"• {nome} ({preco})")
            
            if 'descricao_custom' in item:
                for etapa, brinquedos in item['descricao_custom'].items():
                    if brinquedos: 
                        brinquedos_str = ", ".join(brinquedos)
                        itens_formatados.append(f"  └ {etapa}: {brinquedos_str}")
                        
        return "\n".join(itens_formatados)
        
    except Exception as e:
        print(f"Erro ao formatar JSON de itens: {e}")
        return itens_json 

def formatar_planilha(ws):
    """Aplica formatação de cabeçalho e ajusta colunas."""
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True) 
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # --- (ATUALIZADO) Adicionada Coluna 'O' para Status ---
    col_widths = {
        'A': 8,   # ID BD
        'B': 25,  # ID Google
        'C': 12,  # Data Evento
        'D': 10,  # Horário Evento
        'E': 30,  # Nome Cliente
        'F': 15,  # CPF
        'G': 50,  # Endereço
        'H': 40,  # Itens (Formatados)
        'I': 15,  # Faturamento
        'J': 15,  # Custo Op.
        'K': 15,  # Distância (km)
        'L': 15,  # Custo Combustível
        'M': 15,  # Frete Pago
        'N': 15,  # Lucro (FÓRMULA)
        'O': 15   # Status Pagamento (NOVO)
    }

    # Aplica estilos ao cabeçalho (Linha 1)
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
        cell = ws[f"{col_letter}1"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Aplica estilos às células de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            
            # Moeda:
            if cell.column_letter in ['I', 'J', 'L', 'M', 'N']: 
                cell.alignment = center_align
                cell.number_format = 'R$ #,##0.00'
            
            # Data: (Formato Brasil)
            elif cell.column_letter == 'C': 
                cell.alignment = center_align
                cell.number_format = 'DD/MM/YYYY' 
            
            # Horário:
            elif cell.column_letter == 'D': 
                cell.alignment = center_align
                cell.number_format = 'HH:MM' 
            
            # CPF:
            elif cell.column_letter == 'F': 
                cell.alignment = center_align
            
            # Distância (km):
            elif cell.column_letter == 'K': 
                cell.alignment = center_align
                cell.number_format = '0,0 "km"'
            
            # (NOVA REGRA) Status Pagamento
            elif cell.column_letter == 'O':
                cell.alignment = center_align

            # Texto (default):
            else: 
                cell.alignment = left_align
                
    ws.freeze_panes = 'A2'
    print("Formatação da planilha concluída.")


def sincronizar_db_para_excel():
    """
    Busca todos os dados, incluindo status de pagamento, e os escreve em um arquivo Excel.
    A coluna 'Lucro' é escrita como uma FÓRMULA do Excel.
    """
    print(f"Iniciando sincronização para {ARQUIVO_EXCEL}...")
    
    try:
        # 1. Buscar dados do DB (agora com status de pagamento)
        vendas_db = database.buscar_todas_vendas()
        
        # ... (Seu código para criar planilha vazia se não houver dados) ...
        if not vendas_db:
            print("Nenhum dado encontrado no banco de dados para sincronizar.")
            # ... (código para criar cabeçalho em planilha vazia) ...
            headers = [
                "ID BD", "ID Google", "Data Evento", "Horário Evento", "Nome Cliente", "CPF", "Endereço",
                "Itens (Formatados)", "Faturamento", "Custo Op.", "Distância (km)", 
                "Custo Combustível", "Frete Pago", "Lucro (Líquido)", "Status Pagamento" # (NOVO)
            ]
            # ... (resto do código de planilha vazia) ...
            return

        # 2. Criar ou Carregar a Planilha
        if os.path.exists(ARQUIVO_EXCEL):
            wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
            if PLANILHA_ABA in wb.sheetnames:
                ws = wb[PLANILHA_ABA]
                ws.delete_rows(1, ws.max_row + 1)
                print(f"Aba '{PLANILHA_ABA}' existente limpa.")
            else:
                ws = wb.create_sheet(PLANILHA_ABA)
                print(f"Aba '{PLANILHA_ABA}' criada.")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = PLANILHA_ABA
            print(f"Novo arquivo '{ARQUIVO_EXCEL}' criado.")

        # 3. Definir Cabeçalhos (Headers) (ATUALIZADO)
        headers = [
            "ID BD", "ID Google", "Data Evento", "Horário Evento", "Nome Cliente", "CPF", "Endereço",
            "Itens (Formatados)", "Faturamento", "Custo Op.", "Distância (km)", 
            "Custo Combustível", "Frete Pago", "Lucro (Líquido)", "Status Pagamento" # (NOVO)
        ]
        ws.append(headers)
        
        # Mapeamento de Colunas do Excel
        # ...
        # N = Lucro (Líquido)
        # O = Status Pagamento (NOVO)
        # -------------------------------------

        # 4. Iterar e Escrever Dados + FÓRMULA
        
        current_row_index = 2 
        
        for venda in vendas_db:
            # venda[0] = id
            # ...
            # venda[13] = frete_valor_pago
            # venda[14] = status_pagamento (NOVO)
            
            
            # --- (Processando os dados antes de inserir) ---
            data_evento_str = venda[2]
            data_evento_obj = None
            if data_evento_str:
                try:
                    data_evento_obj = datetime.strptime(data_evento_str, '%Y-%m-%d').date()
                except ValueError:
                    data_evento_obj = data_evento_str 
            
            cpf_formatado = formatar_cpf(venda[5])
            itens_formatados = formatar_itens_para_excel(venda[7])
            
            row_data = [
                venda[0],  # A: ID BD
                venda[1],  # B: ID Google
                data_evento_obj, # C: Data Evento
                venda[3],  # D: Horário Evento
                venda[4],  # E: Nome Cliente
                cpf_formatado, # F: CPF
                venda[6],  # G: Endereço
                itens_formatados, # H: Itens
                venda[8],  # I: Faturamento
                venda[9],  # J: Custo Op.
                venda[11], # K: Distância (km)
                venda[12], # L: Custo Combustível
                venda[13]  # M: Frete Pago
            ]
            
            # Fórmula de Lucro
            lucro_formula_excel = f"=I{current_row_index}-J{current_row_index}-L{current_row_index}"
            row_data.append(lucro_formula_excel) # N: Lucro (Líquido)

            # --- (NOVO) Adiciona a coluna de status ---
            row_data.append(venda[14]) # O: Status Pagamento

            ws.append(row_data)
            
            current_row_index += 1

        # 5. Aplicar Formatação
        formatar_planilha(ws)

        # 6. Salvar o Arquivo
        wb.save(ARQUIVO_EXCEL)
        print(f"Sincronização concluída com sucesso. {len(vendas_db)} registros salvos.")
        
    except PermissionError:
        print(f"ERRO: Permissão negada. A planilha '{ARQUIVO_EXCEL}' está aberta? Feche-a e tente novamente.")
    except Exception as e:
        print(f"Erro inesperado durante a sincronização do Excel: {e}")

if __name__ == "__main__":
    print("Executando sincronização manual do Excel...")
    database.inicializar_banco() 
    sincronizar_db_para_excel()